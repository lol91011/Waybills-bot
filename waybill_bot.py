import logging
import asyncio
import os
import math
import time
import threading
import sys
from datetime import datetime
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import BufferedInputFile, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from aiogram.utils.keyboard import ReplyKeyboardBuilder
import openpyxl
from openpyxl.styles import Alignment
import requests
from flask import Flask, jsonify

# Определение среды развертывания
IS_RAILWAY = os.getenv('RAILWAY_ENVIRONMENT', '').lower() == 'true'
PORT = int(os.getenv('PORT', 5000))

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('bot.log') if not IS_RAILWAY else logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Инициализация Flask приложения для health checks
app = Flask(__name__)

@app.route('/health')
def health_check():
    return jsonify({'status': 'ok', 'bot': 'running'})

def run_flask():
    app.run(host='0.0.0.0', port=PORT)

# Запуск Flask в отдельном потоке
if IS_RAILWAY:
    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()

# Список рекламных ключевых слов
AD_KEYWORDS = [
    'купить', 'продать', 'скидка', 'акция', 'бесплатно', 
    'реклама', 'подпишись', 'канал', 'промокод', 'распродажа',
    'топовый', 'гарантия', 'доставка', 'магазин', 'товар'
]

# Механизм keep-alive
def keep_alive():
    while True:
        try:
            # Простое действие для поддержания активности
            logger.info("Выполняю keep-alive запрос...")
            bot.get_me()
            time.sleep(300)  # 5 минут
        except Exception as e:
            logger.error(f"Ошибка в keep-alive: {e}")
            time.sleep(60)

# Фильтр рекламы
def is_advertisement(text):
    if not text:
        return False
    text = text.lower()
    return any(keyword in text for keyword in AD_KEYWORDS)

class WaybillStates(StatesGroup):
    waiting_for_number = State()
    waiting_for_driver = State()
    waiting_for_car = State()

# Конфигурация
API_TOKEN = "8066885623:AAH4DKVqNfqx5OSRwT4LZL9Io_CzG2RgaqI"
YANDEX_GEOCODE_API = "5ae1bdae-4867-439c-a12f-5b0e4e7ac859"
ORS_API_KEY = "5b3ce3597851110001cf624891fb1b2e0e1d43aa89e9147212dc82c1"

# Получаем путь к папке, где находится скрипт
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "Шаблон_Путевой_лист_Форма3.xlsx")

# Проверка существования шаблона при старте
if not os.path.exists(TEMPLATE_PATH):
    logger.error(f"Файл шаблона не найден по пути: {TEMPLATE_PATH}")
    print(f"🛑 ОШИБКА: Файл шаблона не найден!")
    print(f"🔍 Полный путь: {TEMPLATE_PATH}")
    print("ℹ Убедитесь, что файл 'Шаблон_Путевой_лист_Форма3.xlsx' находится в той же папке, что и скрипт бота")
    exit(1)

# Инициализация бота
bot = Bot(token=API_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

# Хранилище данных пользователей
user_data = {}

# Состояния бота
class Form(StatesGroup):
    driver_info = State()
    date = State()
    start_mileage = State()
    start_fuel = State()
    route_address = State()
    confirm_address = State()
    fuel_consumption = State()
    end_mileage = State()
    fuel_refill = State()
    refill_amount = State()
    confirm_report = State()

def make_start_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="/start")]],
        resize_keyboard=True
    )

def make_yes_no_keyboard():
    builder = ReplyKeyboardBuilder()
    builder.add(KeyboardButton(text="Да"))
    builder.add(KeyboardButton(text="Нет"))
    builder.adjust(2)
    return builder.as_markup(resize_keyboard=True)

def make_done_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="🏁 /готово")]],
        resize_keyboard=True
    )

def make_fuel_refill_keyboard():
    builder = ReplyKeyboardBuilder()
    builder.add(KeyboardButton(text="⛽ Заправиться"))
    builder.add(KeyboardButton(text="➡ Продолжить без заправки"))
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True)

def make_date_keyboard():
    builder = ReplyKeyboardBuilder()
    builder.add(KeyboardButton(text="📅 Сегодня"))
    builder.add(KeyboardButton(text="✏ Ввести вручную"))
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True)

def make_fuel_options_keyboard(last_fuel: float):
    builder = ReplyKeyboardBuilder()
    builder.add(KeyboardButton(text=f"Использовать остаток: {last_fuel:.3f} л"))
    builder.add(KeyboardButton(text="Ввести новое значение"))
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True)

def shorten_address(full_address: str) -> str:
    """Сокращает полный адрес до улицы и номера дома"""
    parts = [part.strip() for part in full_address.split(',')]
    if len(parts) <= 1:
        return full_address
    
    street_part = parts[-2] if len(parts) >= 2 else ''
    house_part = parts[-1] if len(parts) >= 1 else ''
    
    shortened = f"{street_part}, {house_part}" if street_part and house_part else full_address
    
    replacements = {
        "улица": "ул.",
        "проспект": "пр.",
        "проезд": "пр-д",
        "переулок": "пер.",
        "бульвар": "б-р",
        "шоссе": "ш.",
        "набережная": "наб.",
        "дом": "д.",
        "строение": "стр.",
        "корпус": "к."
    }
    
    for full, short in replacements.items():
        shortened = shortened.replace(f" {full} ", f" {short} ")
        shortened = shortened.replace(f" {full},", f" {short},")
    
    return shortened

async def geocode_address(address: str) -> tuple:
    """Геокодирование адреса"""
    try:
        url = "https://geocode-maps.yandex.ru/1.x/"
        params = {
            "apikey": YANDEX_GEOCODE_API,
            "geocode": f"Волгоградская область, {address}",
            "format": "json",
        }
        
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        if data["response"]["GeoObjectCollection"]["featureMember"]:
            geo = data["response"]["GeoObjectCollection"]["featureMember"][0]["GeoObject"]
            full_address = geo["metaDataProperty"]["GeocoderMetaData"]["text"]
            coords = list(map(float, geo["Point"]["pos"].split()))
            return full_address, coords, None
        
        params["geocode"] = address
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        if not data["response"]["GeoObjectCollection"]["featureMember"]:
            return None, None, "Адрес не найден"
            
        geo = data["response"]["GeoObjectCollection"]["featureMember"][0]["GeoObject"]
        full_address = geo["metaDataProperty"]["GeocoderMetaData"]["text"]
        coords = list(map(float, geo["Point"]["pos"].split()))
        
        return full_address, coords, None
        
    except Exception as e:
        logger.error(f"Ошибка геокодирования: {e}")
        return None, None, "Ошибка при поиске адреса"

async def calculate_exact_distance(start_coords: list, end_coords: list) -> float:
    """Расчет точного расстояния по дорогам"""
    try:
        headers = {"Authorization": ORS_API_KEY}
        params = {
            "start": f"{start_coords[0]},{start_coords[1]}",
            "end": f"{end_coords[0]},{end_coords[1]}"
        }
        
        response = requests.get(
            "https://api.openrouteservice.org/v2/directions/driving-car",
            headers=headers,
            params=params,
            timeout=10
        )
        response.raise_for_status()
        data = response.json()
        
        return data["features"][0]["properties"]["segments"][0]["distance"] / 1000
        
    except Exception as e:
        logger.error(f"Ошибка расчета расстояния: {e}")
        # При ошибке API используем упрощенный расчет
        simplified_dist = math.sqrt((end_coords[0]-start_coords[0])**2 + (end_coords[1]-start_coords[1])**2) * 111
        return simplified_dist

def adjust_segments(exact_segments: list, total_actual_distance: int) -> list:
    """Подгоняет отрезки так, чтобы каждый был кратен 10 км с сохранением пропорций"""
    # Рассчитываем суммарное точное расстояние
    total_exact_distance = sum(exact_segments)
    
    # Рассчитываем коэффициенты для каждого отрезка
    coefficients = [seg / total_exact_distance for seg in exact_segments]
    
    # Распределяем общее расстояние с учетом кратности 10 км
    remaining_distance = total_actual_distance
    adjusted_segments = []
    
    for i, coeff in enumerate(coefficients):
        if i == len(coefficients) - 1:
            # Для последнего отрезка используем оставшееся расстояние
            adjusted_segments.append(remaining_distance)
        else:
            # Рассчитываем предполагаемое расстояние для отрезка
            proposed = round((total_actual_distance * coeff) / 10) * 10
            proposed = max(10, proposed)  # Минимум 10 км
            
            # Проверяем, чтобы не превысить оставшееся расстояние
            if proposed > remaining_distance:
                proposed = remaining_distance
            
            adjusted_segments.append(proposed)
            remaining_distance -= proposed
    
    return adjusted_segments

@dp.message(Command("start", "help"))
async def cmd_start(message: types.Message, state: FSMContext):
    """Обработчик команд /start и /help"""
    try:
        user_id = str(message.from_user.id)
        
        welcome_msg = (
            "🚗 *Добро пожаловать в бота для заполнения путевых листов!*\n\n"
            "Я помогу вам быстро и удобно:\n"
            "• Создать путевой лист 📝\n"
            "• Рассчитать пробег и расход топлива ⛽\n"
            "• Сформировать отчет в Excel 📊\n\n"
            "Давайте начнем!"
        )
        
        if user_id in user_data:
            await message.answer(
                "📋 *Текущие данные:*\n"
                f"👤 Водитель: {user_data[user_id].get('driver_name', 'не указано')}\n"
                f"🚘 Марка: {user_data[user_id].get('car_model', 'не указано')}\n"
                f"🔢 Госномер: {user_data[user_id].get('car_number', 'не указано')}\n\n"
                "Хотите изменить данные?",
                reply_markup=make_yes_no_keyboard(),
                parse_mode="Markdown"
            )
        else:
            user_data[user_id] = {}
            await message.answer(
                welcome_msg + "\n\n"
                "📝 *Введите ФИО водителя:*",
                reply_markup=ReplyKeyboardRemove(),
                parse_mode="Markdown"
            )
        await state.set_state(Form.driver_info)
    except Exception as e:
        logger.error(f"Error in cmd_start: {e}")
        await message.answer(
            "⚠ Произошла ошибка. Пожалуйста, попробуйте еще раз.",
            reply_markup=ReplyKeyboardRemove()
        )

@dp.message(Form.driver_info)
async def process_driver_info(message: types.Message, state: FSMContext):
    """Обработка данных водителя"""
    try:
        user_id = str(message.from_user.id)
        text = message.text.lower()
        
        if text in ['нет', 'no'] and user_id in user_data:
            await message.answer(
                "📅 Введите дату путевого листа или нажмите 'Сегодня':",
                reply_markup=make_date_keyboard()
            )
            await state.set_state(Form.date)
            return
        
        if 'driver_name' not in user_data[user_id]:
            user_data[user_id]['driver_name'] = message.text
            await message.answer("🚘 Введите марку машины:")
            return
        
        if 'car_model' not in user_data[user_id]:
            user_data[user_id]['car_model'] = message.text
            await message.answer("🔢 Введите госномер машины:")
            return
        
        user_data[user_id]['car_number'] = message.text
        await message.answer(
            "📅 Введите дату путевого листа или нажмите 'Сегодня':",
            reply_markup=make_date_keyboard()
        )
        await state.set_state(Form.date)
    except Exception as e:
        logger.error(f"Ошибка в process_driver_info: {e}")
        await message.answer(
            "⚠ Произошла ошибка. Пожалуйста, начните снова с команды /start.",
            reply_markup=ReplyKeyboardRemove()
        )

@dp.message(Form.date)
async def process_date(message: types.Message, state: FSMContext):
    """Обработка даты"""
    try:
        user_id = str(message.from_user.id)
        
        if message.text == "📅 Сегодня":
            today = datetime.now().strftime("%d.%m.%Y")
            user_data[user_id]['date'] = today
        elif message.text == "✏ Ввести вручную":
            await message.answer("📅 Введите дату в формате ДД.ММ.ГГГГ:", reply_markup=ReplyKeyboardRemove())
            return
        else:
            try:
                datetime.strptime(message.text, "%d.%m.%Y")
                user_data[user_id]['date'] = message.text
            except ValueError:
                await message.answer("❌ Неверный формат даты. Введите в формате ДД.ММ.ГГГГ:")
                return
        
        last_fuel = user_data[user_id].get('last_fuel', None)
        if last_fuel is not None:
            await message.answer(
                f"⛽ Остаток топлива с прошлой поездки: {last_fuel:.3f} л",
                reply_markup=make_fuel_options_keyboard(last_fuel)
            )
        else:
            await message.answer("⛽ Введите начальное количество топлива (л):", reply_markup=ReplyKeyboardRemove())
        
        await state.set_state(Form.start_fuel)
    except Exception as e:
        logger.error(f"Ошибка в process_date: {e}")
        await message.answer(
            "⚠ Произошла ошибка. Пожалуйста, введите дату еще раз.",
            reply_markup=ReplyKeyboardRemove()
        )

@dp.message(Form.start_fuel)
async def process_start_fuel(message: types.Message, state: FSMContext):
    """Обработка начального количества топлива"""
    try:
        user_id = str(message.from_user.id)
        
        if 'last_fuel' in user_data[user_id]:
            if message.text.startswith("Использовать остаток:"):
                # Извлекаем точное значение из текста кнопки
                last_fuel_str = message.text.split(":")[1].strip().replace(" л", "")
                try:
                    last_fuel = float(last_fuel_str.replace(",", "."))
                    user_data[user_id]['start_fuel'] = last_fuel
                    await state.set_state(Form.start_mileage)
                    await message.answer("🛣️ Введите начальный пробег (км):", reply_markup=ReplyKeyboardRemove())
                    return
                except ValueError:
                    await message.answer("❌ Ошибка при обработке остатка топлива. Введите значение вручную:")
                    return
            elif message.text == "Ввести новое значение":
                await message.answer("⛽ Введите начальное количество топлива (л):", reply_markup=ReplyKeyboardRemove())
                return
        
        try:
            fuel = float(message.text.replace(",", "."))
            if fuel <= 0:
                await message.answer("❌ Количество топлива должно быть больше 0. Введите корректное значение:")
                return
                
            user_data[user_id]['start_fuel'] = fuel
            await state.set_state(Form.start_mileage)
            await message.answer("🛣️ Введите начальный пробег (км):")
        except ValueError:
            await message.answer("❌ Некорректное значение. Введите число:")
            
    except Exception as e:
        logger.error(f"Ошибка в process_start_fuel: {e}")
        await message.answer(
            "⚠ Произошла ошибка. Пожалуйста, введите данные еще раз.",
            reply_markup=ReplyKeyboardRemove()
        )

@dp.message(Form.start_mileage)
async def process_start_mileage(message: types.Message, state: FSMContext):
    """Обработка начального пробега"""
    try:
        user_id = str(message.from_user.id)
        
        if not message.text.isdigit():
            await message.answer("❌ Введите число!")
            return
        
        user_data[user_id]['start_mileage'] = int(message.text)
        user_data[user_id]['route'] = {'points': [], 'exact_segments': [], 'adjusted_segments': [], 'total_distance': 0}
        
        start_point = user_data[user_id].get('next_start_point', None)
        if start_point:
            await message.answer(
                f"📍 Последняя точка из предыдущей поездки: {shorten_address(start_point)}\n"
                "Использовать ее как точку отправления?",
                reply_markup=make_yes_no_keyboard()
            )
            await state.set_state(Form.route_address)
        else:
            await message.answer(
                "📍 Введите первый адрес маршрута (точка отправления):",
                reply_markup=ReplyKeyboardRemove()
            )
            await state.set_state(Form.route_address)
    except Exception as e:
        logger.error(f"Ошибка в process_start_mileage: {e}")
        await message.answer(
            "⚠ Произошла ошибка. Пожалуйста, введите пробег еще раз.",
            reply_markup=ReplyKeyboardRemove()
        )

@dp.message(Form.route_address)
async def process_route_address(message: types.Message, state: FSMContext):
    """Обработка адресов маршрута"""
    try:
        user_id = str(message.from_user.id)
        text = message.text.lower()
        
        # Проверяем, есть ли сохраненная точка и это первый адрес в маршруте
        if 'next_start_point' in user_data[user_id] and len(user_data[user_id]['route']['points']) == 0:
            if text in ['да', 'yes']:
                # Используем сохраненную точку
                address = user_data[user_id]['next_start_point']
                # Получаем координаты для сохраненного адреса
                full_address, coords, error = await geocode_address(address)
                if error:
                    await message.answer(f"❌ Ошибка: {error}")
                    return
                
                user_data[user_id]['route']['points'].append({'address': full_address, 'coords': coords})
                await message.answer(
                    f"📍 Точка отправления: {shorten_address(full_address)}\n"
                    "📍 Введите следующий адрес или /готово:",
                    reply_markup=make_done_keyboard()
                )
                await state.set_state(Form.route_address)
                return
            elif text in ['нет', 'no']:
                del user_data[user_id]['next_start_point']
                await message.answer("📍 Введите адрес отправления:", reply_markup=ReplyKeyboardRemove())
                return
            else:
                await message.answer("Пожалуйста, используйте кнопки 'Да' или 'Нет'")
                return
        
        if message.text == "🏁 /готово":
            if len(user_data[user_id]['route']['points']) < 2:
                await message.answer("❌ Нужно минимум 2 адреса в маршруте!")
                return
            
            if len(user_data[user_id]['route']['points']) > 0:
                # Сохраняем последнюю точку для следующей поездки
                user_data[user_id]['next_start_point'] = user_data[user_id]['route']['points'][-1]['address']
            
            await state.set_state(Form.fuel_consumption)
            await message.answer(
                f"🛣️ Маршрут построен. Точное расстояние: {sum(user_data[user_id]['route']['exact_segments']):.1f} км\n"
                "⛽ Введите средний расход топлива на 100 км (л):",
                reply_markup=ReplyKeyboardRemove()
            )
            return
        
        # Обычная обработка ввода адреса
        address, coords, error = await geocode_address(message.text)
        
        if error:
            await message.answer(error)
            return
        if not address:
            await message.answer("❌ Адрес не найден. Уточните название:")
            return
        
        await state.update_data(temp_address=address, temp_coords=coords)
        await message.answer(
            f"📍 Найден адрес: {shorten_address(address)}\n"
            "Подтверждаете?",
            reply_markup=make_yes_no_keyboard()
        )
        await state.set_state(Form.confirm_address)
    except Exception as e:
        logger.error(f"Ошибка в process_route_address: {e}")
        await message.answer(
            "⚠ Произошла ошибка. Пожалуйста, введите адрес еще раз.",
            reply_markup=ReplyKeyboardRemove()
        )

@dp.message(Form.confirm_address)
async def process_confirm_address(message: types.Message, state: FSMContext):
    """Подтверждение адреса"""
    try:
        user_id = str(message.from_user.id)
        text = message.text.lower()
        
        if text not in ['да', 'нет']:
            await message.answer("Пожалуйста, используйте кнопки 'Да' или 'Нет'")
            return
        
        if text == 'нет':
            await state.set_state(Form.route_address)
            await message.answer("📍 Введите адрес еще раз:", reply_markup=ReplyKeyboardRemove())
            return
        
        data = await state.get_data()
        address = data['temp_address']
        coords = data['temp_coords']
        
        route = user_data[user_id]['route']
        route['points'].append({'address': address, 'coords': coords})
        
        if len(route['points']) == 1:
            await message.answer(
                f"📍 Точка отправления: {shorten_address(address)}\n"
                "📍 Введите следующий адрес или /готово:",
                reply_markup=make_done_keyboard()
            )
        else:
            prev_coords = route['points'][-2]['coords']
            exact_distance = await calculate_exact_distance(prev_coords, coords)
            route['exact_segments'].append(exact_distance)
            
            await message.answer(
                f"📍 Адрес добавлен: {shorten_address(address)}\n"
                f"🛣️ Точное расстояние: {exact_distance:.1f} км\n"
                f"📊 Суммарное точное расстояние: {sum(route['exact_segments']):.1f} км\n"
                "📍 Введите следующий адрес или /готово:",
                reply_markup=make_done_keyboard()
            )
        
        await state.set_state(Form.route_address)
    except Exception as e:
        logger.error(f"Ошибка в process_confirm_address: {e}")
        await message.answer(
            "⚠ Произошла ошибка. Пожалуйста, введите адрес еще раз.",
            reply_markup=ReplyKeyboardRemove()
        )

@dp.message(Form.fuel_consumption)
async def process_fuel_consumption(message: types.Message, state: FSMContext):
    """Обработка расхода топлива"""
    try:
        user_id = str(message.from_user.id)
        
        try:
            fuel_consumption = float(message.text.replace(",", "."))
            if fuel_consumption <= 0:
                raise ValueError("Расход должен быть больше 0")
            user_data[user_id]['fuel_per_100km'] = fuel_consumption
        except ValueError:
            await message.answer("❌ Некорректное значение. Введите положительное число:")
            return
        
        await state.set_state(Form.end_mileage)
        await message.answer("🛣️ Введите конечный пробег (км):")
    except Exception as e:
        logger.error(f"Ошибка в process_fuel_consumption: {e}")
        await message.answer(
            "⚠ Произошла ошибка. Пожалуйста, введите расход еще раз.",
            reply_markup=ReplyKeyboardRemove()
        )

@dp.message(Form.end_mileage)
async def process_end_mileage(message: types.Message, state: FSMContext):
    """Обработка конечного пробега"""
    try:
        user_id = str(message.from_user.id)
        
        if not message.text.isdigit():
            await message.answer("❌ Введите число!")
            return
        
        end_mileage = int(message.text)
        start_mileage = user_data[user_id]['start_mileage']
        
        if end_mileage <= start_mileage:
            await message.answer("❌ Конечный пробег должен быть больше начального!")
            return
        
        actual_distance = end_mileage - start_mileage
        exact_distance = sum(user_data[user_id]['route']['exact_segments'])
        fuel_per_100km = user_data[user_id]['fuel_per_100km']
        start_fuel = user_data[user_id]['start_fuel']
        
        # Подгоняем отрезки под фактический пробег с кратностью 10 км
        adjusted_segments = adjust_segments(
            user_data[user_id]['route']['exact_segments'],
            actual_distance
        )
        
        user_data[user_id]['route']['adjusted_segments'] = adjusted_segments
        user_data[user_id]['route']['total_distance'] = sum(adjusted_segments)
        user_data[user_id]['actual_distance'] = actual_distance
        
        fuel_used = (actual_distance * fuel_per_100km) / 100
        end_fuel = start_fuel - fuel_used
        
        # Сохраняем данные о расходе
        user_data[user_id]['fuel_used_actual'] = fuel_used
        user_data[user_id]['fuel_used_from_tank'] = min(fuel_used, start_fuel)
        user_data[user_id]['fuel_deficit'] = max(0, fuel_used - start_fuel)
        user_data[user_id]['end_fuel_unrefilled'] = end_fuel
        user_data[user_id]['end_mileage'] = end_mileage
        user_data[user_id]['actual_distance'] = actual_distance
        
        route_text = "🛣️ *Маршрут:*\n"
        points = user_data[user_id]['route']['points']
        exact_segments = user_data[user_id]['route']['exact_segments']
        adjusted_segments = user_data[user_id]['route']['adjusted_segments']
        
        for i in range(1, len(points)):
            from_addr = shorten_address(points[i-1]['address'])
            to_addr = shorten_address(points[i]['address'])
            route_text += f"{i}. {from_addr} → {to_addr} ({adjusted_segments[i-1]} км, точное: {exact_segments[i-1]:.1f} км)\n"
        
        fuel_warning = ""
        if end_fuel < 0:
            fuel_warning = "\n⚠ Внимание! Топлива израсходовано больше, чем было в баке!"
        
        await message.answer(
            "📋 *Проверьте данные:*\n\n"
            f"📅 Дата: {user_data[user_id]['date']}\n"
            f"👤 Водитель: {user_data[user_id]['driver_name']}\n"
            f"🚘 Автомобиль: {user_data[user_id]['car_model']} ({user_data[user_id]['car_number']})\n\n"
            f"🛣️ Пробег: {start_mileage} → {end_mileage} км (всего {actual_distance} км)\n"
            f"⛽ Расход топлива: {fuel_per_100km:.3f} л/100км\n"
            f"⛽ Израсходовано всего: {fuel_used:.3f} л\n"
            f"⛽ Из них из бака: {min(fuel_used, start_fuel):.3f} л\n"
            f"⛽ Дефицит топлива: {max(0, fuel_used - start_fuel):.3f} л\n"
            f"⛽ Остаток в баке: {max(0, end_fuel):.3f} л\n"
            f"{fuel_warning}\n"
            f"{route_text}\n"
            "Хотите заправиться?",
            reply_markup=make_fuel_refill_keyboard(),
            parse_mode="Markdown"
        )
        await state.set_state(Form.fuel_refill)
    except Exception as e:
        logger.error(f"Ошибка в process_end_mileage: {e}")
        await message.answer(
            "⚠ Произошла ошибка. Пожалуйста, введите пробег еще раз.",
            reply_markup=ReplyKeyboardRemove()
        )

@dp.message(Form.fuel_refill)
async def process_fuel_refill(message: types.Message, state: FSMContext):
    """Обработка заправки"""
    try:
        user_id = str(message.from_user.id)
        
        if message.text == "⛽ Заправиться":
            await message.answer("⛽ Введите количество заправленного топлива (л):", reply_markup=ReplyKeyboardRemove())
            await state.set_state(Form.refill_amount)
        elif message.text == "➡ Продолжить без заправки":
            # Без заправки - сохраняем фактический остаток (не может быть отрицательным)
            user_data[user_id]['last_fuel'] = max(0, user_data[user_id]['end_fuel_unrefilled'])
            user_data[user_id]['fuel_refill'] = 0
            await show_report_confirmation(message, user_id)
            await state.set_state(Form.confirm_report)
        else:
            await message.answer("Пожалуйста, используйте кнопки", reply_markup=make_fuel_refill_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в process_fuel_refill: {e}")
        await message.answer(
            "⚠ Произошла ошибка. Пожалуйста, попробуйте еще раз.",
            reply_markup=ReplyKeyboardRemove()
        )

@dp.message(Form.refill_amount)
async def process_refill_amount(message: types.Message, state: FSMContext):
    """Обработка количества топлива"""
    try:
        user_id = str(message.from_user.id)
        
        try:
            refill = float(message.text.replace(",", "."))
            if refill <= 0:
                raise ValueError("Количество должно быть больше 0")
            
            deficit = user_data[user_id].get('fuel_deficit', 0)
            
            if deficit > 0:
                # Если был дефицит, сначала покрываем его
                if refill <= deficit:
                    # Вся заправка пошла на покрытие дефицита
                    user_data[user_id]['last_fuel'] = 0
                else:
                    # Часть заправки покрыла дефицит, остаток - новый остаток
                    user_data[user_id]['last_fuel'] = refill - deficit
            else:
                # Дефицита не было - просто добавляем к остатку
                user_data[user_id]['last_fuel'] = max(0, user_data[user_id]['end_fuel_unrefilled']) + refill
            
            user_data[user_id]['fuel_refill'] = refill
            
            await message.answer(
                f"⛽ Заправлено {refill:.3f} л.\n"
                f"Дефицит топлива: {deficit:.3f} л\n"
                f"Новый остаток в баке: {user_data[user_id]['last_fuel']:.3f} л",
                reply_markup=ReplyKeyboardRemove()
            )
            await show_report_confirmation(message, user_id)
            await state.set_state(Form.confirm_report)
        except ValueError:
            await message.answer("❌ Некорректное значение. Введите положительное число:")
    except Exception as e:
        logger.error(f"Ошибка в process_refill_amount: {e}")
        await message.answer(
            "⚠ Произошла ошибка. Пожалуйста, введите количество еще раз.",
            reply_markup=ReplyKeyboardRemove()
        )

async def show_report_confirmation(message: types.Message, user_id: str):
    """Подтверждение отчета"""
    try:
        data = user_data[user_id]
        route = data['route']
        
        route_text = "🛣️ *Маршрут:*\n"
        points = route['points']
        exact_segments = route['exact_segments']
        adjusted_segments = route['adjusted_segments']
        
        for i in range(1, len(points)):
            from_addr = shorten_address(points[i-1]['address'])
            to_addr = shorten_address(points[i]['address'])
            route_text += f"{i}. {from_addr} → {to_addr} ({adjusted_segments[i-1]} км, точное: {exact_segments[i-1]:.1f} км)\n"
        
        refill_info = f"\n⛽ Заправлено топлива: {data.get('fuel_refill', 0):.3f} л\n" if 'fuel_refill' in data else ""
        deficit_info = f"\n⛽ Дефицит топлива: {data.get('fuel_deficit', 0):.3f} л\n" if data.get('fuel_deficit', 0) > 0 else ""
        
        await message.answer(
            "📋 *Итоговые данные:*\n\n"
            f"📅 Дата: {data['date']}\n"
            f"👤 Водитель: {data['driver_name']}\n"
            f"🚘 Автомобиль: {data['car_model']} ({data['car_number']})\n\n"
            f"🛣️ Пробег: {data['start_mileage']} → {data['end_mileage']} км (всего {data['actual_distance']} км)\n"
            f"⛽ Расход топлива: {data['fuel_per_100km']:.3f} л/100км\n"
            f"⛽ Израсходовано всего: {data['fuel_used_actual']:.3f} л\n"
            f"⛽ Из них из бака: {data['fuel_used_from_tank']:.3f} л\n"
            f"{deficit_info}"
            f"⛽ Остаток топлива: {data['last_fuel']:.3f} л\n"
            f"{refill_info}\n"
            f"{route_text}\n"
            "Сформировать путевой лист?",
            reply_markup=make_yes_no_keyboard(),
            parse_mode="Markdown"
        )
    except Exception as e:
        logger.error(f"Ошибка в show_report_confirmation: {e}")
        await message.answer(
            "⚠ Произошла ошибка при формировании отчета.",
            reply_markup=ReplyKeyboardRemove()
        )

@dp.message(Form.confirm_report)
async def process_confirm_report(message: types.Message, state: FSMContext):
    """Финальное подтверждение"""
    try:
        user_id = str(message.from_user.id)
        text = message.text.lower()
        
        if text not in ['да', 'нет']:
            await message.answer("Пожалуйста, используйте кнопки 'Да' или 'Нет'")
            return
        
        if text == 'нет':
            await message.answer("❌ Данные не сохранены. Начните заново с /start", reply_markup=make_start_keyboard())
            await state.clear()
            return
        
        await generate_report(message, user_id)
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка в process_confirm_report: {e}")
        await message.answer(
            "⚠ Произошла ошибка. Пожалуйста, попробуйте еще раз.",
            reply_markup=ReplyKeyboardRemove()
        )

async def generate_report(message: types.Message, user_id: str):
    """Генерация Excel-отчета по шаблону"""
    try:
        data = user_data[user_id]
        
        # Загрузка шаблона
        try:
            wb = openpyxl.load_workbook(TEMPLATE_PATH)
            ws = wb.worksheets[0]  # Первый лист
            ws2 = wb.worksheets[1]  # Второй лист
        except Exception as e:
            error_msg = (
                "⚠ *Ошибка при загрузке шаблона!*\n\n"
                f"Проверьте, что файл шаблона:\n"
                f"1. Находится в папке с ботом\n"
                f"2. Имеет название `Шаблон_Путевой_лист_Форма3.xlsx`\n"
                f"3. Не поврежден\n\n"
                f"Техническая информация: `{str(e)}`"
            )
            await message.answer(error_msg, parse_mode="Markdown")
            return

        # Заполнение данных
        try:
            # Форматирование даты
            day, month, year = data['date'].split('.')
            month_name = {
                '01': 'января', '02': 'февраля', '03': 'марта', '04': 'апреля',
                '05': 'мая', '06': 'июня', '07': 'июля', '08': 'августа',
                '09': 'сентября', '10': 'октября', '11': 'ноября', '12': 'декабря'
            }[month]
            
            # Вставка даты в ячейку D5
            ws['D5'] = f'Срок действия с «{day}» {month_name} {year}г. по «{day}» {month_name} {year}г.'
            
            # Основные данные
            ws['V10'] = data['car_model']    # Марка автомобиля
            ws['AI11'] = data['car_number']  # Госномер
            ws['M12'] = data['driver_name']  # ФИО водителя
            
            # Топливо и пробег
            ws['BT34'] = data.get('fuel_refill', 0)  # Заправлено
            ws['BT37'] = data['start_fuel']          # При выезде
            ws['BT38'] = data['last_fuel']           # При возвращении
            
            # Расход топлива (дублируем в обе ячейки)
            fuel_consumption = data['fuel_per_100km']
            ws['BT39'] = fuel_consumption  # Норма
            ws['BT40'] = fuel_consumption  # Фактический (дублируем то же значение)
            
            # Пробег
            ws['BU19'] = data['start_mileage']  # Начальный
            ws['BT45'] = data['end_mileage']   # Конечный

            # Маршрут (второй лист)
            total_distance = 0
            for i, segment in enumerate(data['route']['adjusted_segments'][:25], start=1):
                row = 5 + i - 1
                from_addr = shorten_address(data['route']['points'][i-1]['address'])
                to_addr = shorten_address(data['route']['points'][i]['address'])
                
                ws2[f'E{row}'] = from_addr
                ws2[f'H{row}'] = to_addr
                ws2[f'O{row}'] = segment
                total_distance += segment

            ws2['E35'] = total_distance  # Общее расстояние

        except Exception as e:
            logger.error(f"Ошибка при заполнении данных: {e}")
            await message.answer(
                "⚠ Произошла ошибка при заполнении данных в шаблоне.\n"
                f"Ошибка: `{str(e)}`",
                parse_mode="Markdown"
            )
            return

        # Сохранение и отправка файла
        filename = f"Путевой лист {data['date']}.xlsx"
        try:
            wb.save(filename)
            
            with open(filename, 'rb') as file:
                success_msg = (
                    "✅ *Путевой лист успешно сформирован!*\n\n"
                    f"📅 Дата: {data['date']}\n"
                    f"👤 Водитель: {data['driver_name']}\n"
                    f"🚘 Автомобиль: {data['car_model']} ({data['car_number']})\n"
                    f"🛣️ Пробег: {data['start_mileage']} → {data['end_mileage']} км\n"
                    f"⛽ Расход топлива: {data['fuel_per_100km']:.3f} л/100км\n"
                    f"⛽ Израсходовано: {data['fuel_used_actual']:.3f} л\n"
                    f"⛽ Остаток топлива: {data['last_fuel']:.3f} л"
                )
                
                await message.answer_document(
                    document=BufferedInputFile(file.read(), filename=filename),
                    caption=success_msg,
                    parse_mode="Markdown"
                )
            
            os.remove(filename)
            
            # Добавляем кнопку "Старт" после успешного создания отчета
            await message.answer(
                "Хотите создать новый путевой лист?",
                reply_markup=make_start_keyboard()
            )
            
        except Exception as e:
            logger.error(f"Ошибка при сохранении/отправке файла: {e}")
            await message.answer(
                "⚠ Произошла ошибка при сохранении или отправке файла.\n"
                f"Ошибка: `{str(e)}`",
                parse_mode="Markdown"
            )

    except Exception as e:
        logger.error(f"Критическая ошибка при генерации отчета: {e}", exc_info=True)
        await message.answer(
            "⚠ *Критическая ошибка!*\n\n"
            "При создании отчета произошла непредвиденная ошибка.\n"
            "Пожалуйста, попробуйте еще раз или обратитесь к администратору.\n\n"
            f"Техническая информация: `{str(e)}`",
            parse_mode="Markdown"
        )

@dp.message(Command("debug"))
async def cmd_debug(message: types.Message):
    """Команда для отладки"""
    try:
        current_dir = os.path.abspath(os.path.dirname(__file__))
        files = "\n".join(os.listdir(current_dir))
        template_status = "✅ найден" if os.path.exists(TEMPLATE_PATH) else "❌ НЕ найден"
        
        debug_msg = (
            "🛠 *Отладочная информация:*\n\n"
            f"📂 Текущая директория: `{current_dir}`\n"
            f"📄 Статус шаблона: {template_status}\n"
            f"🔍 Полный путь к шаблону: `{TEMPLATE_PATH}`\n\n"
            "📋 Содержимое директории:\n"
            f"```\n{files}\n```"
        )
        
        await message.answer(
            debug_msg,
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardRemove()
        )
    except Exception as e:
        logger.error(f"Ошибка в cmd_debug: {e}")
        await message.answer(
            "⚠ Произошла ошибка при получении отладочной информации.",
            reply_markup=ReplyKeyboardRemove()
        )

@dp.message()
async def handle_unexpected_messages(message: types.Message):
    """Обработка непредусмотренных сообщений"""
    # Проверка на рекламу
    if is_advertisement(message.text):
        logger.info(f"Обнаружено рекламное сообщение от {message.from_user.id}: {message.text}")
        try:
            await message.delete()
            await message.answer("Реклама запрещена в этом чате.")
        except Exception as e:
            logger.error(f"Ошибка при удалении рекламного сообщения: {e}")
        return
    
    await message.answer(
        "❌ Я не понимаю это сообщение. Пожалуйста, используйте команды:\n"
        "/start - начать заполнение путевого листа\n"
        "/help - получить помощь\n"
        "/debug - отладочная информация",
        reply_markup=make_start_keyboard()
    )

async def on_startup(bot: Bot):
    """Действия при запуске бота"""
    try:
        await bot.delete_webhook(drop_pending_updates=True)
        logger.info("Бот успешно запущен")
        print("🤖 Бот успешно запущен")
        
        # Запуск keep-alive потока
        keep_alive_thread = threading.Thread(target=keep_alive, daemon=True)
        keep_alive_thread.start()
        
        if IS_RAILWAY:
            logger.info("Работает в среде Railway")
    except Exception as e:
        logger.error(f"Ошибка при запуске бота: {e}")
        print(f"🛑 Ошибка при запуске бота: {e}")

async def on_shutdown(bot: Bot):
    """Действия при остановке бота"""
    try:
        logger.info("Завершение работы бота...")
        await bot.session.close()
        logger.info("Бот успешно остановлен")
    except Exception as e:
        logger.error(f"Ошибка при остановке бота: {e}")

# ... (весь предыдущий код остается без изменений до функции main)

async def main():
    """Запуск бота"""
    try:
        logger.info("Запуск бота...")
        print("🤖 Бот запускается...")

        # Удаляем вебхук и ждем завершения
        await bot.delete_webhook(drop_pending_updates=True)
        await asyncio.sleep(1)

        # Запускаем поллинг без указания allowed_updates
        await dp.start_polling(
            bot,
            on_startup=on_startup,
            on_shutdown=on_shutdown,
            skip_updates=True
        )
    except Exception as e:
        logger.critical(f"Ошибка: {e}")
        print(f"💥 Критическая ошибка: {e}")
        time.sleep(10)
        os.execv(sys.executable, ['python'] + sys.argv)
    finally:
        await bot.session.close()

if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Бот остановлен пользователем")
        print("\n🛑 Бот остановлен")
    except Exception as e:
        logger.critical(f"Необработанное исключение: {e}")
        print(f"💥 Критическая ошибка: {e}")
        # Попытка автоматического перезапуска
        time.sleep(10)
        os.execv(sys.executable, ['python'] + sys.argv)