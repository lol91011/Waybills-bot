import logging
import os
from collections import defaultdict
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (ApplicationBuilder, CommandHandler, MessageHandler,
                          ConversationHandler, ContextTypes, filters)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
def get_start_keyboard():
    return ReplyKeyboardMarkup(
        [[KeyboardButton("Сегодня")], [KeyboardButton("Ввести вручную")]],
        resize_keyboard=True, one_time_keyboard=True
    )

def get_confirmation_keyboard():
    return ReplyKeyboardMarkup(
        [[KeyboardButton("Да"), KeyboardButton("Нет")]],
        resize_keyboard=True, one_time_keyboard=True
    )

import requests

YANDEX_API_KEY = "0a947709-13f1-4dee-9ce9-74ea971cffb0"
ORS_API_KEY = "5b3ce3597851110001cf624891fb1b2e0e1d43aa89e9147212dc82c1"

def geocode_yandex(address):
    url = "https://geocode-maps.yandex.ru/1.x/"
    params = {
        "apikey": YANDEX_API_KEY,
        "geocode": f"Волгоградская область, Россия, {address}",
        "format": "json",
        "lang": "ru_RU"
    }
    resp = requests.get(url, params=params).json()
    try:
        pos = resp["response"]["GeoObjectCollection"]["featureMember"][0]["GeoObject"]["Point"]["pos"]
        lon, lat = pos.split()
        return float(lat), float(lon)
    except:
        return None

def get_route_distance(start_coords, end_coords):
    url = "https://api.openrouteservice.org/v2/directions/driving-car"
    headers = {"Authorization": ORS_API_KEY}
    payload = {
        "coordinates": [[start_coords[1], start_coords[0]], [end_coords[1], end_coords[0]]]
    }
    resp = requests.post(url, json=payload, headers=headers)
    try:
        data = resp.json()
        meters = data["features"][0]["properties"]["segments"][0]["distance"]
        return round(meters / 1000, 1)
    except:
        return None


(ASK_DATE, ASK_NAME, ASK_CAR, ASK_START_ODOMETER, ASK_FUEL_CONSUMPTION,
 ASK_FUEL_BEFORE_TRIP, ASK_ROUTE, ASK_END_ODOMETER, CONFIRMATION) = range(9)

user_data_store = defaultdict(dict)
frequent_addresses = defaultdict(lambda: {"Туркменская, 14а": 1})

def get_date_keyboard():
    return ReplyKeyboardMarkup([[KeyboardButton("Сегодня")]], resize_keyboard=True)

def get_address_keyboard(user_id):
    addresses = sorted(frequent_addresses[user_id].items(), key=lambda x: -x[1])
    keyboard = [[KeyboardButton(addr[0])] for addr in addresses[:3]]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Выберите дату:", reply_markup=get_start_keyboard())
    return ASK_DATE

async def ask_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    context.user_data["date"] = update.message.text
    if "name" in user_data_store[user_id]:
        return await ask_car(update, context)
    await update.message.reply_text("Введите ФИО водителя:")
    return ASK_NAME

async def ask_car(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if "name" not in user_data_store[user_id]:
        user_data_store[user_id]["name"] = update.message.text
    if "car" in user_data_store[user_id]:
        return await ask_start_odometer(update, context)
    await update.message.reply_text("Введите марку автомобиля и номер:")
    return ASK_CAR

async def ask_start_odometer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if "car" not in user_data_store[user_id]:
        user_data_store[user_id]["car"] = update.message.text
    await update.message.reply_text("Введите начальный пробег:")
    return ASK_START_ODOMETER

async def ask_fuel_consumption(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["start_odometer"] = update.message.text
    await update.message.reply_text("Введите средний расход топлива на 100 км:")
    return ASK_FUEL_CONSUMPTION

async def ask_fuel_before_trip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["fuel_before"] = update.message.text
    await update.message.reply_text("Введите количество топлива до выезда:")
    return ASK_FUEL_BEFORE_TRIP

async def ask_route(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["fuel_before"] = update.message.text
    user_id = update.effective_user.id
    await update.message.reply_text("Введите маршрут:", reply_markup=get_address_keyboard(user_id))
    return ASK_ROUTE

async def ask_end_odometer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["route"] = update.message.text
    user_id = update.effective_user.id
    frequent_addresses[user_id][update.message.text] += 1
    await update.message.reply_text("Введите конечный пробег:")
    return ASK_END_ODOMETER

async def confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["end_odometer"] = update.message.text
    user_id = update.effective_user.id
    summary = ""
    summary += f"Дата: {context.user_data['date']}\n"
    summary += f"Водитель: {user_data_store[user_id]['name']}\n"
    summary += f"Машина: {user_data_store[user_id]['car']}\n"
    summary += f"Начальный пробег: {context.user_data['start_odometer']}\n"
    summary += f"Средний расход: {context.user_data['consumption']}\n"
    summary += f"Топливо до выезда: {context.user_data['fuel_before']}\n"
    summary += f"Маршрут: {context.user_data.get('route_text', '')}\n"
    summary += f"Конечный пробег: {context.user_data['end_odometer']}"
    await update.message.reply_text(summary)
    filename = generate_waybill_excel(user_id, {
        "name": user_data_store[user_id]["name"],
        "car": user_data_store[user_id]["car"],
        "start_odometer": context.user_data["start_odometer"],
        "end_odometer": context.user_data["end_odometer"],
        "fuel_before": context.user_data["fuel_before"],
        "consumption": context.user_data["consumption"],
        "parsed_routes": context.user_data.get("parsed_routes", [])
    })
    with open(filename, "rb") as doc:
        await update.message.reply_document(doc, filename=filename)
    return ConversationHandler.END


async def confirm_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text = update.message.text.lower()
    if text == 'да':
        name = context.user_data['last_address']
        coords = context.user_data['last_coords']
        points = context.user_data['route_points']
        if points:
            prev_coords = points[-1][1]
            dist = await get_distance_ors(prev_coords, coords)
            dist_rounded = int(round(dist / 10.0) * 10)
            if 'route_text' not in context.user_data:
                context.user_data['route_text'] = ''
            context.user_data['route_text'] += f"{points[-1][0]} → {name}: {dist_rounded} км ({dist} км)\n"
        points.append((name, coords))
        await update.message.reply_text("Добавьте следующий адрес или напишите 'готово'")
        return ASK_ROUTE
    else:
        await update.message.reply_text("Введите адрес снова:")
        return ASK_ROUTE
def main():
    token = os.getenv("BOT_TOKEN")
    app = ApplicationBuilder().token(token).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            ASK_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_name)],
            ASK_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_car)],
            ASK_CAR: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_start_odometer)],
            ASK_START_ODOMETER: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_fuel_consumption)],
            ASK_FUEL_CONSUMPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_fuel_before_trip)],
            ASK_FUEL_BEFORE_TRIP: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_route)],
            ASK_ROUTE: [
                MessageHandler(filters.Regex("^(да|нет)$"), confirm_address),
                MessageHandler(filters.TEXT & ~filters.COMMAND, ask_route)
            ],
            ASK_END_ODOMETER: [MessageHandler(filters.TEXT & ~filters.COMMAND, confirmation)],
        },
        fallbacks=[]
    )

    app.add_handler(conv_handler)
    app.run_polling()

if __name__ == "__main__":
    main()


from openpyxl import load_workbook
from datetime import datetime

def generate_waybill_excel(user_id, user_data):
    template_path = "plist_template.xlsx"
    date_str = datetime.now().strftime("%Y-%m-%d")
    output_filename = f"waybill_{user_id}_{date_str}.xlsx"

    wb = load_workbook(template_path)
    ws = wb.active

    ws["M12"] = user_data["name"]
    ws["BU19"] = int(user_data["start_odometer"])
    ws["BT45"] = int(user_data["end_odometer"])
    ws["BT37"] = float(user_data["fuel_before"])
    ws["BT38"] = float(user_data["fuel_after"]) if "fuel_after" in user_data else 0
    ws["BT39"] = float(user_data["consumption"])
    ws["V10"]  = user_data["car"].split()[0]
    ws["AL11"] = user_data["car"]

    start_row = 5
    for i, entry in enumerate(user_data.get("parsed_routes", [])):
        ws[f"E{start_row + i}"] = entry["from"]
        ws[f"H{start_row + i}"] = entry["to"]
        ws[f"O{start_row + i}"] = entry["rounded_km"]

    wb.save(output_filename)
    return output_filename